#!/usr/bin/env python3
"""
Bid Leveling Workbook Generator
Creates a professional multi-worksheet Excel workbook from structured bid data.
Usage: Import and call create_workbook(bid_data, output_path)
       Or run directly for a sample/test output.
"""
import os, sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles ──
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
SUBHEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
SUBHEADER_FONT = Font(name='Arial', bold=True, color='1F4E79', size=10)
DATA_FONT = Font(name='Arial', size=10)
TOTAL_FONT = Font(name='Arial', bold=True, size=11)
TOTAL_FILL = PatternFill('solid', fgColor='E2EFDA')
WARNING_FILL = PatternFill('solid', fgColor='FCE4EC')
WARNING_FONT = Font(name='Arial', italic=True, color='CC0000', size=10)
BLUE_INPUT = Font(name='Arial', color='0000FF', size=10)
CURRENCY_FMT = '$#,##0'
CURRENCY_DET = '$#,##0.00'
PERCENT_FMT = '0.0%'
SCORE_FMT = '0.0'
BORDER = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))


def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = BORDER


def auto_width(ws, mn=12, mx=40):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max(max((len(str(c.value or "")) for c in col), default=0) + 2, mn), mx)
        ws.column_dimensions[letter].width = width


def create_workbook(bid_data, output_path):
    wb = Workbook()
    bidders = bid_data["bidders"]
    n = len(bidders)
    names = [b["name"] for b in bidders]
    proj = bid_data.get("project", {})

    # ── Tab 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1F4E79"
    r = 1
    ws.cell(r, 1, "BID LEVELING SUMMARY").font = Font(name='Arial', bold=True, size=16, color='1F4E79')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2 + n)
    r += 2
    for lbl, val in [("Project Name", proj.get("name", "")), ("Project Type", proj.get("type", "")),
                      ("Location", proj.get("location", "")), ("Square Footage", proj.get("sf", "")),
                      ("Date", proj.get("date", "")), ("Delivery Method", proj.get("delivery_method", ""))]:
        ws.cell(r, 1, lbl).font = SUBHEADER_FONT
        ws.cell(r, 2, val).font = DATA_FONT
        r += 1
    r += 1
    ws.cell(r, 1, "RECOMMENDATION").font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    r += 1
    ws.cell(r, 1, bid_data.get("recommendation", "")).font = DATA_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2 + n)
    ws.row_dimensions[r].height = 60
    ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical='top')
    r += 2
    ws.cell(r, 1, "QUICK COMPARISON").font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    r += 1
    hdrs = ["Metric"] + names
    for c, h in enumerate(hdrs, 1): ws.cell(r, c, h)
    style_header(ws, r, len(hdrs))
    r += 1
    ws.cell(r, 1, "Original Bid").font = DATA_FONT
    for i, b in enumerate(bidders):
        cell = ws.cell(r, i+2, b["original_bid"])
        cell.number_format = CURRENCY_FMT; cell.font = DATA_FONT; cell.border = BORDER
    r += 1
    ws.cell(r, 1, "Leveled Bid").font = TOTAL_FONT
    for i in range(n):
        cell = ws.cell(r, i+2)
        cell.value = f"='Leveled Comparison'!{get_column_letter(i+2)}7"
        cell.number_format = CURRENCY_FMT; cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL; cell.border = BORDER
    r += 1
    ws.cell(r, 1, "Rank").font = TOTAL_FONT
    for i in range(n):
        cell = ws.cell(r, i+2)
        cell.value = f"='Ranking Detail'!{get_column_letter(i+3)}9"
        cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL; cell.border = BORDER
    r += 2
    warnings = bid_data.get("data_warnings", [])
    if warnings:
        ws.cell(r, 1, "DATA QUALITY WARNINGS").font = Font(name='Arial', bold=True, size=12, color='CC0000')
        r += 1
        for w in warnings:
            ws.cell(r, 1, f"⚠ {w}").font = WARNING_FONT
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2 + n)
            r += 1
    auto_width(ws)
    ws.column_dimensions['A'].width = 25

    # ── Tab 2: Leveled Comparison ──
    ws2 = wb.create_sheet("Leveled Comparison")
    ws2.sheet_properties.tabColor = "2E75B6"
    hdrs = ["Item"] + names
    for c, h in enumerate(hdrs, 1): ws2.cell(1, c, h)
    style_header(ws2, 1, len(hdrs))
    labels = ["Original Bid", "Net Allowance Adjustments", "Net Exclusion Add-backs", "Risk Premium", "", "Leveled Bid Total", "$/SF", "Risk Score (0-100)"]
    for ri, lbl in enumerate(labels, 2):
        ws2.cell(ri, 1, lbl).font = DATA_FONT; ws2.cell(ri, 1).border = BORDER
    for i, b in enumerate(bidders):
        c = i + 2
        cell = ws2.cell(2, c, b["original_bid"]); cell.number_format = CURRENCY_FMT; cell.font = BLUE_INPUT; cell.border = BORDER
        cell = ws2.cell(3, c); cell.value = f"=SUMIF('Allowance Adjustments'!A:A,\"{names[i]}\",'Allowance Adjustments'!D:D)"; cell.number_format = CURRENCY_FMT; cell.font = DATA_FONT; cell.border = BORDER
        cell = ws2.cell(4, c); cell.value = f"=SUMIF('Exclusion Add-backs'!A:A,\"{names[i]}\",'Exclusion Add-backs'!D:D)"; cell.number_format = CURRENCY_FMT; cell.font = DATA_FONT; cell.border = BORDER
        cell = ws2.cell(5, c); cell.value = f"=SUMIF('Risk Analysis'!A:A,\"{names[i]}\",'Risk Analysis'!H:H)"; cell.number_format = CURRENCY_FMT; cell.font = DATA_FONT; cell.border = BORDER
    ws2.cell(7, 1, "Leveled Bid Total").font = TOTAL_FONT; ws2.cell(7, 1).fill = TOTAL_FILL
    for i in range(n):
        cl = get_column_letter(i+2)
        cell = ws2.cell(7, i+2); cell.value = f"={cl}2+{cl}3+{cl}4+{cl}5"
        cell.number_format = CURRENCY_FMT; cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL; cell.border = BORDER
    sf = proj.get("sf")
    for i in range(n):
        cl = get_column_letter(i+2)
        cell = ws2.cell(8, i+2)
        if sf and str(sf).replace(',','').replace('.','').isdigit():
            cell.value = f"={cl}7/{float(str(sf).replace(',',''))}"; cell.number_format = CURRENCY_DET
        else:
            cell.value = "N/A"
        cell.font = DATA_FONT; cell.border = BORDER
    for i, b in enumerate(bidders):
        risk = next((r for r in bid_data.get("risk_scores", []) if r["bidder"] == b["name"]), None)
        cell = ws2.cell(9, i+2, risk["composite"] if risk else 0)
        cell.number_format = SCORE_FMT; cell.font = DATA_FONT; cell.border = BORDER
    ws2.freeze_panes = "B2"; auto_width(ws2)

    # ── Tab 3: Ranking Detail ──
    ws3 = wb.create_sheet("Ranking Detail")
    ws3.sheet_properties.tabColor = "548235"
    hdrs = ["Category", "Weight"] + names
    for c, h in enumerate(hdrs, 1): ws3.cell(1, c, h)
    style_header(ws3, 1, len(hdrs))
    wts = bid_data.get("ranking_weights", {"cost": 0.45, "scope": 0.20, "risk": 0.15, "schedule": 0.10, "transparency": 0.10})
    cats = [("Leveled Cost", wts.get("cost", 0.45)), ("Scope Completeness", wts.get("scope", 0.20)),
            ("Risk Profile", wts.get("risk", 0.15)), ("Schedule Realism", wts.get("schedule", 0.10)),
            ("Bid Transparency", wts.get("transparency", 0.10))]
    for ri, (cat, wt) in enumerate(cats, 2):
        ws3.cell(ri, 1, cat).font = DATA_FONT; ws3.cell(ri, 1).border = BORDER
        cell = ws3.cell(ri, 2, wt); cell.number_format = PERCENT_FMT; cell.font = BLUE_INPUT; cell.border = BORDER
    ls, le = get_column_letter(2), get_column_letter(1+n)
    for i, b in enumerate(bidders):
        ci = i + 3; tcl = get_column_letter(i+2)
        cell = ws3.cell(2, ci); cell.value = f"=MIN('Leveled Comparison'!{ls}7:{le}7)/'Leveled Comparison'!{tcl}7*100"
        cell.number_format = SCORE_FMT; cell.font = DATA_FONT; cell.border = BORDER
        ranking = next((r for r in bid_data.get("ranking_scores", []) if r["bidder"] == b["name"]), {})
        for j, key in enumerate(["scope_score", "risk_score", "schedule_score", "transparency_score"], 3):
            cell = ws3.cell(j, ci, ranking.get(key, 0))
            cell.number_format = SCORE_FMT; cell.font = BLUE_INPUT; cell.border = BORDER
    ws3.cell(8, 1, "Weighted Total").font = TOTAL_FONT; ws3.cell(8, 1).fill = TOTAL_FILL; ws3.cell(8, 1).border = BORDER
    ws3.cell(8, 2).fill = TOTAL_FILL; ws3.cell(8, 2).border = BORDER
    for i in range(n):
        cl = get_column_letter(i+3)
        cell = ws3.cell(8, i+3); cell.value = f"=SUMPRODUCT($B$2:$B$6,{cl}2:{cl}6)"
        cell.number_format = SCORE_FMT; cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL; cell.border = BORDER
    ws3.cell(9, 1, "Rank").font = TOTAL_FONT; ws3.cell(9, 1).fill = TOTAL_FILL; ws3.cell(9, 1).border = BORDER
    ws3.cell(9, 2).fill = TOTAL_FILL; ws3.cell(9, 2).border = BORDER
    rs, re_ = get_column_letter(3), get_column_letter(2+n)
    for i in range(n):
        cl = get_column_letter(i+3)
        cell = ws3.cell(9, i+3); cell.value = f"=RANK({cl}8,${rs}$8:${re_}$8)"
        cell.number_format = '0'; cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL; cell.border = BORDER
    ws3.freeze_panes = "C2"; auto_width(ws3)

    # ── Tab 4: Allowance Adjustments ──
    ws4 = wb.create_sheet("Allowance Adjustments")
    ws4.sheet_properties.tabColor = "BF8F00"
    for c, h in enumerate(["Bidder", "Category", "Original Allowance", "Net Adjustment", "Standardized Allowance", "Notes"], 1):
        ws4.cell(1, c, h)
    style_header(ws4, 1, 6)
    for ri, adj in enumerate(bid_data.get("allowance_adjustments", []), 2):
        ws4.cell(ri, 1, adj.get("bidder", "")).font = DATA_FONT
        ws4.cell(ri, 2, adj.get("category", "")).font = DATA_FONT
        ws4.cell(ri, 3, adj.get("original", 0)).number_format = CURRENCY_FMT
        ws4.cell(ri, 3).font = DATA_FONT
        ws4.cell(ri, 4, adj.get("net", 0)).number_format = CURRENCY_FMT
        ws4.cell(ri, 4).font = DATA_FONT
        ws4.cell(ri, 5, adj.get("standardized", 0)).number_format = CURRENCY_FMT
        ws4.cell(ri, 5).font = DATA_FONT
        ws4.cell(ri, 6, adj.get("notes", "")).font = DATA_FONT
        for c in range(1, 7): ws4.cell(ri, c).border = BORDER
    ws4.freeze_panes = "A2"; auto_width(ws4)

    # ── Tab 5: Exclusion Add-backs ──
    ws5 = wb.create_sheet("Exclusion Add-backs")
    ws5.sheet_properties.tabColor = "C55A11"
    for c, h in enumerate(["Bidder", "Item", "Required/Optional", "Estimated Cost", "Basis"], 1):
        ws5.cell(1, c, h)
    style_header(ws5, 1, 5)
    for ri, exc in enumerate(bid_data.get("exclusion_addbacks", []), 2):
        ws5.cell(ri, 1, exc.get("bidder", "")).font = DATA_FONT
        ws5.cell(ri, 2, exc.get("item", "")).font = DATA_FONT
        ws5.cell(ri, 3, exc.get("required", "Required")).font = DATA_FONT
        ws5.cell(ri, 4, exc.get("cost", 0)).number_format = CURRENCY_FMT
        ws5.cell(ri, 4).font = DATA_FONT
        ws5.cell(ri, 5, exc.get("basis", "")).font = DATA_FONT
        for c in range(1, 6): ws5.cell(ri, c).border = BORDER
    ws5.freeze_panes = "A2"; auto_width(ws5)

    # ── Tab 6: Risk Analysis ──
    ws6 = wb.create_sheet("Risk Analysis")
    ws6.sheet_properties.tabColor = "CC0000"
    risk_hdrs = ["Bidder", "Scope Gaps (0-20)", "Allowance Realism (0-20)", "Schedule Risk (0-20)",
                 "Pricing Balance (0-20)", "Track Record (0-20)", "Composite (0-100)",
                 "Risk Premium ($)", "Risk Premium (%)", "Justification"]
    for c, h in enumerate(risk_hdrs, 1): ws6.cell(1, c, h)
    style_header(ws6, 1, len(risk_hdrs))
    for ri, risk in enumerate(bid_data.get("risk_scores", []), 2):
        ws6.cell(ri, 1, risk.get("bidder", "")).font = DATA_FONT
        for j, k in enumerate(["scope_gaps", "allowance_realism", "schedule_risk", "pricing_balance", "track_record"], 2):
            ws6.cell(ri, j, risk.get(k, 0)).font = BLUE_INPUT; ws6.cell(ri, j).border = BORDER
        cell = ws6.cell(ri, 7); cell.value = f"=SUM(B{ri}:F{ri})"; cell.number_format = SCORE_FMT; cell.font = TOTAL_FONT; cell.border = BORDER
        ws6.cell(ri, 8, risk.get("premium_dollar", 0)).number_format = CURRENCY_FMT; ws6.cell(ri, 8).font = DATA_FONT; ws6.cell(ri, 8).border = BORDER
        ws6.cell(ri, 9, risk.get("premium_pct", 0)).number_format = PERCENT_FMT; ws6.cell(ri, 9).font = DATA_FONT; ws6.cell(ri, 9).border = BORDER
        ws6.cell(ri, 10, risk.get("justifications", "")).font = DATA_FONT; ws6.cell(ri, 10).alignment = Alignment(wrap_text=True); ws6.cell(ri, 10).border = BORDER
    ws6.freeze_panes = "B2"; auto_width(ws6); ws6.column_dimensions['J'].width = 50

    # ── Tab 7: Assumptions & Gaps ──
    ws7 = wb.create_sheet("Assumptions & Gaps")
    ws7.sheet_properties.tabColor = "7F7F7F"
    ws7.cell(1, 1, "Category"); ws7.cell(1, 2, "Detail")
    style_header(ws7, 1, 2)
    r = 2
    ws7.cell(r, 1, "ASSUMPTIONS").font = SUBHEADER_FONT; ws7.cell(r, 1).fill = SUBHEADER_FILL
    ws7.cell(r, 2).fill = SUBHEADER_FILL; r += 1
    for a in bid_data.get("assumptions", []):
        ws7.cell(r, 1, "Assumption").font = DATA_FONT; ws7.cell(r, 2, a).font = DATA_FONT
        ws7.cell(r, 2).alignment = Alignment(wrap_text=True)
        ws7.cell(r, 1).border = BORDER; ws7.cell(r, 2).border = BORDER; r += 1
    r += 1
    ws7.cell(r, 1, "DATA QUALITY WARNINGS").font = SUBHEADER_FONT; ws7.cell(r, 1).fill = WARNING_FILL
    ws7.cell(r, 2).fill = WARNING_FILL; r += 1
    for w in bid_data.get("data_warnings", []):
        ws7.cell(r, 1, "Warning").font = WARNING_FONT; ws7.cell(r, 1).fill = WARNING_FILL; ws7.cell(r, 1).border = BORDER
        ws7.cell(r, 2, w).font = WARNING_FONT; ws7.cell(r, 2).fill = WARNING_FILL; ws7.cell(r, 2).border = BORDER
        ws7.cell(r, 2).alignment = Alignment(wrap_text=True); r += 1
    ws7.column_dimensions['A'].width = 25; ws7.column_dimensions['B'].width = 80

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    # Run with sample data for testing
    sample = {
        "project": {"name": "Sample Office Building", "type": "Commercial", "location": "Nashville, TN", "sf": "50000", "date": "2025-01-15", "delivery_method": "Hard Bid"},
        "bidders": [{"name": "Contractor A", "original_bid": 8500000}, {"name": "Contractor B", "original_bid": 7800000}, {"name": "Contractor C", "original_bid": 9200000}],
        "allowance_adjustments": [
            {"bidder": "Contractor A", "category": "Site Conditions", "original": 100000, "standardized": 170000, "net": 70000, "notes": "Below baseline"},
            {"bidder": "Contractor B", "category": "Site Conditions", "original": 50000, "standardized": 170000, "net": 120000, "notes": "Significantly below baseline"},
            {"bidder": "Contractor B", "category": "Material Escalation", "original": 0, "standardized": 234000, "net": 234000, "notes": "Missing entirely"}],
        "exclusion_addbacks": [
            {"bidder": "Contractor B", "item": "Fire Protection", "required": "Required", "cost": 275000, "basis": "$5.50/SF x 50,000 SF"},
            {"bidder": "Contractor B", "item": "Winter Conditions", "required": "Required", "cost": 78000, "basis": "1% of hard cost"}],
        "risk_scores": [
            {"bidder": "Contractor A", "scope_gaps": 5, "allowance_realism": 8, "schedule_risk": 6, "pricing_balance": 4, "track_record": 5, "composite": 28, "premium_pct": 0.02, "premium_dollar": 170000, "justifications": "Moderate risk. Below-baseline allowances offset by solid scope coverage."},
            {"bidder": "Contractor B", "scope_gaps": 14, "allowance_realism": 16, "schedule_risk": 10, "pricing_balance": 8, "track_record": 10, "composite": 58, "premium_pct": 0.04, "premium_dollar": 312000, "justifications": "High risk. Major exclusions, aggressive schedule, missing allowances."},
            {"bidder": "Contractor C", "scope_gaps": 3, "allowance_realism": 4, "schedule_risk": 5, "pricing_balance": 3, "track_record": 5, "composite": 20, "premium_pct": 0.01, "premium_dollar": 92000, "justifications": "Low risk. Most complete bid. Conservative timeline."}],
        "ranking_weights": {"cost": 0.45, "scope": 0.20, "risk": 0.15, "schedule": 0.10, "transparency": 0.10},
        "ranking_scores": [
            {"bidder": "Contractor A", "scope_score": 75, "risk_score": 72, "schedule_score": 80, "transparency_score": 85},
            {"bidder": "Contractor B", "scope_score": 45, "risk_score": 42, "schedule_score": 55, "transparency_score": 50},
            {"bidder": "Contractor C", "scope_score": 90, "risk_score": 80, "schedule_score": 90, "transparency_score": 92}],
        "assumptions": ["Project SF: 50,000 (from bid docs)", "Baseline allowances use Mid-range commercial benchmarks", "Fire protection at $5.50/SF from comparable projects"],
        "data_warnings": ["Contractor B PDF had poorly formatted tables — some line items may have been misread"],
        "recommendation": "Contractor C is recommended as best value despite highest original bid. After leveling, Contractor B's cost advantage disappears due to major exclusions and aggressive allowances. Contractor A is a solid second choice with moderate risk."
    }
    out = sys.argv[1] if len(sys.argv) > 1 else "/tmp/test_bid_leveling.xlsx"
    print(f"Workbook created: {create_workbook(sample, out)}")
