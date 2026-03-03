#!/usr/bin/env python3
import os,sys,json,re
try:
    import pdfplumber
except ImportError:
    os.system(sys.executable+" -m pip install pdfplumber")
    import pdfplumber
try:
    from openpyxl import load_workbook
except ImportError:
    os.system(sys.executable+" -m pip install openpyxl")
    from openpyxl import load_workbook

def extract_pdf(fp):
    r={"filename":os.path.basename(fp),"type":"pdf","text":"","tables":[],"warnings":[]}
    try:
        with pdfplumber.open(fp) as pdf:
            for i,page in enumerate(pdf.pages):
                t=page.extract_text() or ""
                r["text"]+=t
                for j,table in enumerate(page.extract_tables()):
                    if table: r["tables"].append({"page":i+1,"rows":[[str(c).strip() if c else "" for c in row] for row in table]})
    except Exception as e: r["warnings"].append(str(e))
    return r

def extract_excel(fp):
    r={"filename":os.path.basename(fp),"type":"excel","sheets":{},"warnings":[]}
    try:
        wb=load_workbook(fp,data_only=True)
        for name in wb.sheetnames:
            r["sheets"][name]=[[str(c) if c else "" for c in row] for row in wb[name].iter_rows(values_only=True)]
    except Exception as e: r["warnings"].append(str(e))
    return r

def process(d):
    results={"documents":[],"summary":{}}
    for fn in sorted(os.listdir(d)):
        fp=os.path.join(d,fn)
        ext=os.path.splitext(fn)[1].lower()
        if ext==".pdf": results["documents"].append(extract_pdf(fp))
        elif ext in (".xlsx",".xls",".xlsm"): results["documents"].append(extract_excel(fp))
    return results

if __name__=="__main__":
    print(json.dumps(process(sys.argv[1] if len(sys.argv)>1 else "."),indent=2,default=str))
