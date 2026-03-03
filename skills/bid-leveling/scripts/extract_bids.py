#!/usr/bin/env python3
"""
Bid Document Extractor
Reads PDF and Excel bid documents and extracts structured cost data.
Usage: python extract_bids.py [upload_directory]
"""
import os, sys, json, re
import pdfplumber
from openpyxl import load_workbook


def extract_pdf(filepath):
    result = {"filename": os.path.basename(filepath), "type": "pdf", "text": "", "tables": [], "warnings": []}
    try:
        with pdfplumber.open(filepath) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text() or ""
                result["text"] += f"\n--- Page {i+1} ---\n{text}"
                for j, table in enumerate(page.extract_tables()):
                    if table:
                        cleaned = [[str(c).strip() if c else "" for c in row] for row in table]
                        result["tables"].append({"page": i+1, "table_index": j, "rows": cleaned})
            if not result["text"].strip():
                result["warnings"].append("PDF text extraction returned empty — may be scanned/image-based.")
            if not result["tables"]:
                result["warnings"].append("No tables detected in PDF — cost data may be embedded in prose.")
    except Exception as e:
        result["warnings"].append(f"PDF extraction error: {e}")
    result["dollar_amounts"] = extract_numbers(result["text"])
    return result


def extract_excel(filepath):
    result = {"filename": os.path.basename(filepath), "type": "excel", "sheets": {}, "warnings": []}
    try:
        wb = load_workbook(filepath, data_only=True)
        for name in wb.sheetnames:
            rows = [[str(c) if c is not None else "" for c in row] for row in wb[name].iter_rows(values_only=True)]
            result["sheets"][name] = rows
            if not any(any(c.strip() for c in r) for r in rows):
                result["warnings"].append(f"Sheet '{name}' appears empty.")
    except Exception as e:
        result["warnings"].append(f"Excel extraction error: {e}")
    return result


def extract_numbers(text):
    amounts = []
    for match in re.finditer(r'(\$[\d,]+(?:\.\d{2})?)', text):
        raw = match.group(1).replace('$', '').replace(',', '')
        try:
            val = float(raw)
            start = max(0, match.start() - 80)
            context = text[start:match.start()].strip()
            label = context.split('\n')[-1].strip() if context else "Unknown"
            amounts.append({"label": label, "amount": val, "raw": match.group(0)})
        except ValueError:
            pass
    return amounts


def process_uploads(upload_dir):
    results = {"documents": [], "summary": {"pdf_count": 0, "excel_count": 0, "warnings": []}}
    if not os.path.exists(upload_dir):
        results["summary"]["warnings"].append(f"Directory not found: {upload_dir}")
        return results
    for filename in sorted(os.listdir(upload_dir)):
        filepath = os.path.join(upload_dir, filename)
        ext = os.path.splitext(filename)[1].lower()
        if ext == '.pdf':
            results["documents"].append(extract_pdf(filepath))
            results["summary"]["pdf_count"] += 1
        elif ext in ('.xlsx', '.xls', '.xlsm'):
            results["documents"].append(extract_excel(filepath))
            results["summary"]["excel_count"] += 1
    results["summary"]["total_documents"] = len(results["documents"])
    for doc in results["documents"]:
        for w in doc.get("warnings", []):
            results["summary"]["warnings"].append(f"{doc['filename']}: {w}")
    return results


if __name__ == "__main__":
    upload_dir = sys.argv[1] if len(sys.argv) > 1 else "."
    print(json.dumps(process_uploads(upload_dir), indent=2, default=str))
